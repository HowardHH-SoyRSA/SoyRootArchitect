from __future__ import annotations

from dataclasses import replace
from datetime import datetime, timezone
import json
import math
from pathlib import Path
import shutil
import xml.etree.ElementTree as ET

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .geometry import resample_polyline
from .io import write_labeled_ply, write_point_cloud
from .topology import hierarchy_frame, write_editable_hierarchy
from .traits import angle_vectors_frame, lateral_counts_frame, trait_summary_frame
from .types import Normalization, RootPath, TopologyReport


SEGMENT_COLORS = {
    "unassigned": np.array([0.55, 0.55, 0.55]),
    "uncertain": np.array([0.98, 0.48, 0.05]),
    "primary": np.array([0.05, 0.23, 0.88]),
    "order_1": np.array([1.0, 0.0, 1.0]),  # #FF00FF
    "order_2": np.array([0.0, 158.0 / 255.0, 115.0 / 255.0]),  # #009E73
    "order_3": np.array([0.55, 0.20, 0.82]),
    "higher_order": np.array([0.95, 0.65, 0.08]),
}


def order_color(order: int) -> np.ndarray:
    if order <= 0:
        return SEGMENT_COLORS["primary"]
    return SEGMENT_COLORS.get(f"order_{order}", SEGMENT_COLORS["higher_order"])


def export_results(
    output_dir: str | Path,
    original_points: np.ndarray,
    primary_path: np.ndarray,
    lateral_paths: list[RootPath],
    primary_mask: np.ndarray,
    lateral_labels: np.ndarray,
    traits: pd.DataFrame,
    normalization: Normalization,
    metadata: dict,
    *,
    full_points: np.ndarray | None = None,
    triangles: np.ndarray | None = None,
    full_root_labels: np.ndarray | None = None,
    topology_report: TopologyReport | None = None,
) -> None:
    """Write measurement, topology, provenance, figure-input, and PLY outputs."""

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_dir = output_dir / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)
    original_primary_path = normalization.inverse_points(primary_path)
    _write_skeleton_csv(
        output_dir / "primary_skeleton.csv",
        "primary",
        original_primary_path,
    )
    lateral_skeletons = _lateral_skeleton_frame(lateral_paths, normalization)
    lateral_skeletons.to_csv(output_dir / "lateral_skeletons.csv", index=False)

    tables = _trait_tables(traits, lateral_paths, normalization, primary_path)
    traits.to_csv(output_dir / "root_traits.csv", index=False)
    for filename, frame in tables.items():
        frame.to_csv(csv_dir / filename, index=False)
    _write_trait_workbook(output_dir / "traits.xlsx", traits, tables)

    geometry_points = np.asarray(full_points if full_points is not None else original_points, dtype=float)
    geometry_labels = (
        np.asarray(full_root_labels, dtype=int)
        if full_root_labels is not None
        else _analysis_root_labels(primary_mask, lateral_labels)
    )
    if len(geometry_labels) != len(geometry_points):
        raise ValueError("Full-resolution root labels do not match export vertices.")
    colors, root_orders, assignment_states = _label_properties(geometry_labels, lateral_paths)
    write_labeled_ply(
        output_dir / "segmented_root_structure.ply",
        geometry_points,
        triangles=triangles,
        colors=colors,
        root_ids=geometry_labels,
        root_orders=root_orders,
        assignment_states=assignment_states,
    )
    # Compatibility point-cloud output used by the original MVP and simple PLY
    # viewers that ignore custom scalar properties.
    write_point_cloud(output_dir / "segmented_points.ply", geometry_points, colors=colors)
    _write_class_point_clouds(output_dir, geometry_points, geometry_labels, colors)
    skeleton_overlay_layout = _write_skeleton_overlay(
        output_dir / "skeleton_original_overlay.ply",
        geometry_points,
        triangles,
        primary_path,
        lateral_paths,
        normalization,
        metadata,
    )

    editable_paths = [
        replace(
            root,
            points=normalization.inverse_points(root.points),
            insertion_point=(
                None
                if root.insertion_point is None
                else normalization.inverse_points(np.asarray(root.insertion_point)[None, :])[0]
            ),
        )
        for root in lateral_paths
    ]
    primary_trait = traits.loc[traits["root_id"] == "primary"]
    primary_confidence = float(primary_trait.iloc[0]["confidence"]) if len(primary_trait) else 1.0
    primary_qc_flags = (
        [flag for flag in str(primary_trait.iloc[0]["qc_flags"]).split(";") if flag]
        if len(primary_trait)
        else []
    )
    write_editable_hierarchy(
        output_dir / "root_hierarchy.json",
        original_primary_path,
        editable_paths,
        primary_confidence=primary_confidence,
        primary_qc_flags=primary_qc_flags,
    )
    write_rsml(
        output_dir / "root_system.rsml",
        original_primary_path,
        editable_paths,
        traits,
        metadata,
    )
    metadata = dict(metadata)
    metadata["system_summary"] = traits.attrs.get("system_summary", {})
    metadata["root_label_map"] = tables["root_label_map.csv"].to_dict(orient="records")
    metadata["skeleton_original_overlay_layout"] = skeleton_overlay_layout
    if topology_report is not None:
        metadata["topology_report"] = _json_safe(topology_report.__dict__)
    metadata["outputs"] = sorted(
        str(path.relative_to(output_dir)).replace("\\", "/")
        for path in output_dir.rglob("*")
        if path.is_file()
    )
    with (output_dir / "metadata.json").open("w", encoding="utf-8") as handle:
        json.dump(_json_safe(metadata), handle, indent=2, ensure_ascii=False, allow_nan=False)


def write_rsml(
    path: str | Path,
    primary_path: np.ndarray,
    lateral_paths: list[RootPath],
    traits: pd.DataFrame,
    metadata: dict,
) -> Path:
    """Export a hierarchy-preserving RSML 1.x document in mesh units."""

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rsml = ET.Element("rsml")
    metadata_element = ET.SubElement(rsml, "metadata")
    ET.SubElement(metadata_element, "version").text = "1"
    ET.SubElement(metadata_element, "unit").text = "mesh_unit"
    ET.SubElement(metadata_element, "resolution").text = "1"
    ET.SubElement(metadata_element, "last-modified").text = datetime.now(timezone.utc).isoformat()
    ET.SubElement(metadata_element, "software").text = "SoyRootBio"
    ET.SubElement(metadata_element, "user").text = "SoyRootBio user"
    ET.SubElement(metadata_element, "file-key").text = Path(str(metadata.get("source", "root"))).stem
    property_definitions = ET.SubElement(metadata_element, "property-definitions")
    for name, value_type, unit in (
        ("root_order", "integer", "none"),
        ("confidence", "real", "none"),
        ("length", "real", "mesh_unit"),
        ("mean_diameter", "real", "mesh_unit"),
        ("tortuosity", "real", "none"),
    ):
        definition = ET.SubElement(property_definitions, "property-definition")
        ET.SubElement(definition, "label").text = name
        ET.SubElement(definition, "type").text = value_type
        ET.SubElement(definition, "unit").text = unit
    scene = ET.SubElement(rsml, "scene")
    plant = ET.SubElement(scene, "plant", {"id": "plant-1", "label": Path(str(metadata.get("source", "root"))).stem})
    trait_lookup = {str(row["root_id"]): row for _, row in traits.iterrows()}
    children: dict[str, list[RootPath]] = {}
    for root in lateral_paths:
        children.setdefault(root.parent_id, []).append(root)

    primary_element = _rsml_root_element(
        plant,
        root_id="primary",
        label="primary",
        order=0,
        points=np.asarray(primary_path, dtype=float),
        confidence=float(trait_lookup.get("primary", {}).get("confidence", 1.0)),
        trait=trait_lookup.get("primary"),
    )

    def append_children(parent_element: ET.Element, parent_id: str) -> None:
        for child in sorted(children.get(parent_id, []), key=lambda item: item.root_id):
            child_element = _rsml_root_element(
                parent_element,
                root_id=child.root_id,
                label=f"order {child.order}",
                order=int(child.order),
                points=np.asarray(child.points, dtype=float),
                confidence=float(child.confidence),
                trait=trait_lookup.get(child.root_id),
            )
            append_children(child_element, child.root_id)

    append_children(primary_element, "primary")
    tree = ET.ElementTree(rsml)
    ET.indent(tree, space="  ")
    tree.write(path, encoding="utf-8", xml_declaration=True)
    return path


def _rsml_root_element(
    parent: ET.Element,
    *,
    root_id: str,
    label: str,
    order: int,
    points: np.ndarray,
    confidence: float,
    trait,
) -> ET.Element:
    root = ET.SubElement(parent, "root", {"id": root_id, "label": label})
    properties = ET.SubElement(root, "properties")
    values = {
        "root_order": order,
        "confidence": confidence,
        "length": np.nan if trait is None else trait.get("length", np.nan),
        "mean_diameter": np.nan if trait is None else trait.get("mean_diameter", np.nan),
        "tortuosity": np.nan if trait is None else trait.get("tortuosity", np.nan),
    }
    for name, value in values.items():
        ET.SubElement(properties, name, {"value": str(value)})
    geometry = ET.SubElement(root, "geometry")
    polyline = ET.SubElement(geometry, "polyline")
    for point in points:
        ET.SubElement(
            polyline,
            "point",
            {"x": _decimal_text(point[0]), "y": _decimal_text(point[1]), "z": _decimal_text(point[2])},
        )
    return root


def _decimal_text(value: float) -> str:
    """Format an XSD ``decimal`` without forbidden exponent notation."""

    value = float(value)
    if not np.isfinite(value):
        raise ValueError("RSML coordinates must be finite")
    if abs(value) < 5e-16:
        value = 0.0
    return np.format_float_positional(value, precision=15, unique=True, trim="-")


def _trait_tables(
    traits: pd.DataFrame,
    lateral_paths: list[RootPath],
    normalization: Normalization,
    primary_path: np.ndarray,
) -> dict[str, pd.DataFrame]:
    identity = ["root_id", "parent_id", "root_order"]
    length = traits[identity + ["length", "chord_length", "length_unit"]].copy()
    counts = lateral_counts_frame(traits)
    angles = traits.loc[
        traits["root_order"] > 0,
        identity
        + [
            "base_parent_angle_deg",
            "tip_gravity_angle_deg",
            "tip_start_gravity_angle_deg",
            "tip_primary_angle_deg",
        ],
    ].copy()
    tortuosity = traits[identity + ["tortuosity"]].copy()
    surface = traits[identity + ["surface_area", "area_unit", "surface_area_method"]].copy()
    volume = traits[identity + ["volume", "volume_unit", "volume_method"]].copy()
    diameter = traits[
        identity
        + [
            "mean_diameter",
            "median_diameter",
            "minimum_diameter",
            "maximum_diameter",
            "length_unit",
        ]
    ].copy()
    topology = hierarchy_frame(
        lateral_paths,
        normalization=normalization,
        primary_path=primary_path,
        primary_confidence=float(
            traits.loc[traits["root_id"] == "primary", "confidence"].iloc[0]
            if np.any(traits["root_id"] == "primary")
            else 1.0
        ),
        primary_qc_flags=(
            str(traits.loc[traits["root_id"] == "primary", "qc_flags"].iloc[0]).split(";")
            if np.any(traits["root_id"] == "primary")
            else ()
        ),
    )
    label_rows = [
        {
            "numeric_label": -2,
            "root_id": "uncertain",
            "parent_id": "",
            "root_order": "",
            "color_rgb": ",".join(str(int(round(value * 255))) for value in SEGMENT_COLORS["uncertain"]),
        },
        {
            "numeric_label": -1,
            "root_id": "unassigned",
            "parent_id": "",
            "root_order": "",
            "color_rgb": ",".join(str(int(round(value * 255))) for value in SEGMENT_COLORS["unassigned"]),
        },
        {
            "numeric_label": 0,
            "root_id": "primary",
            "parent_id": "",
            "root_order": 0,
            "color_rgb": ",".join(str(int(round(value * 255))) for value in SEGMENT_COLORS["primary"]),
        }
    ]
    for numeric_label, root in enumerate(lateral_paths, start=1):
        label_rows.append(
            {
                "numeric_label": numeric_label,
                "root_id": root.root_id,
                "parent_id": root.parent_id,
                "root_order": int(root.order),
                "color_rgb": ",".join(str(int(round(value * 255))) for value in order_color(root.order)),
            }
        )
    label_map = pd.DataFrame.from_records(label_rows)
    qc = traits[identity + ["confidence", "point_count", "qc_flags"]].copy()
    return {
        "system_summary.csv": trait_summary_frame(traits),
        "root_lengths.csv": length,
        "lateral_counts_by_order.csv": counts,
        "root_angles.csv": angles,
        "root_tortuosity.csv": tortuosity,
        "root_surface_area.csv": surface,
        "root_volume.csv": volume,
        "root_diameter.csv": diameter,
        "angle_vectors.csv": angle_vectors_frame(traits),
        "root_topology.csv": topology,
        "root_qc.csv": qc,
        "root_label_map.csv": label_map,
    }


def _write_trait_workbook(path: Path, traits: pd.DataFrame, tables: dict[str, pd.DataFrame]) -> None:
    sheet_map = {
        "system_summary.csv": "System summary",
        "root_lengths.csv": "Length",
        "lateral_counts_by_order.csv": "Counts by order",
        "root_angles.csv": "Angles",
        "root_tortuosity.csv": "Tortuosity",
        "root_surface_area.csv": "Surface area",
        "root_volume.csv": "Volume",
        "root_diameter.csv": "Diameter",
        "angle_vectors.csv": "Vectors",
        "root_topology.csv": "Topology",
        "root_qc.csv": "QC",
        "root_label_map.csv": "Label map",
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        traits.to_excel(writer, sheet_name="Root traits", index=False)
        for filename, frame in tables.items():
            frame.to_excel(writer, sheet_name=sheet_map[filename], index=False)
        _style_trait_workbook(writer.book)


def _style_trait_workbook(workbook) -> None:
    """Apply compact, audit-friendly navigation and number formatting."""

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        worksheet.row_dimensions[1].height = 32
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        for column_index, cells in enumerate(worksheet.iter_cols(), start=1):
            values = ["" if cell.value is None else str(cell.value) for cell in cells]
            width = min(34, max(10, max((len(value) for value in values), default=0) + 2))
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
            for cell in cells[1:]:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"


def _lateral_skeleton_frame(paths: list[RootPath], normalization: Normalization) -> pd.DataFrame:
    rows = []
    for path in paths:
        original = normalization.inverse_points(path.points)
        for node_id, xyz in enumerate(original):
            rows.append(
                {
                    "root_id": path.root_id,
                    "parent_id": path.parent_id,
                    "root_order": int(path.order),
                    "node_id": node_id,
                    "x": xyz[0],
                    "y": xyz[1],
                    "z": xyz[2],
                    "coordinate_unit": "mesh_unit",
                    "confidence": path.confidence,
                    "qc_flags": ";".join(path.qc_flags),
                }
            )
    return pd.DataFrame(
        rows,
        columns=[
            "root_id",
            "parent_id",
            "root_order",
            "node_id",
            "x",
            "y",
            "z",
            "coordinate_unit",
            "confidence",
            "qc_flags",
        ],
    )


def _write_skeleton_csv(
    path: Path,
    root_id: str,
    points: np.ndarray,
) -> None:
    rows = [
        {
            "root_id": root_id,
            "node_id": i,
            "x": p[0],
            "y": p[1],
            "z": p[2],
            "coordinate_unit": "mesh_unit",
        }
        for i, p in enumerate(points)
    ]
    pd.DataFrame(rows).to_csv(path, index=False)


def _analysis_root_labels(primary_mask: np.ndarray, lateral_labels: np.ndarray) -> np.ndarray:
    labels = np.full(len(primary_mask), -1, dtype=int)
    labels[np.asarray(primary_mask, dtype=bool)] = 0
    lateral_labels = np.asarray(lateral_labels, dtype=int)
    labels[lateral_labels > 0] = lateral_labels[lateral_labels > 0]
    labels[lateral_labels < 0] = -2
    return labels


def _label_properties(labels: np.ndarray, paths: list[RootPath]) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    colors = np.tile(SEGMENT_COLORS["unassigned"], (len(labels), 1))
    orders = np.full(len(labels), 255, dtype=np.uint8)
    states = np.zeros(len(labels), dtype=np.uint8)
    uncertain = labels == -2
    colors[uncertain] = SEGMENT_COLORS["uncertain"]
    orders[uncertain] = 254
    states[uncertain] = 2
    primary = labels == 0
    colors[primary] = SEGMENT_COLORS["primary"]
    orders[primary] = 0
    states[primary] = 1
    for label, path in enumerate(paths, start=1):
        mask = labels == label
        colors[mask] = order_color(path.order)
        orders[mask] = np.uint8(min(253, max(0, int(path.order))))
        states[mask] = 1
    return colors, orders, states


def _write_class_point_clouds(output_dir: Path, points: np.ndarray, labels: np.ndarray, colors: np.ndarray) -> None:
    masks = {
        "primary_points.ply": labels == 0,
        "lateral_points.ply": labels > 0,
        "unassigned_points.ply": labels == -1,
        "uncertain_points.ply": labels == -2,
    }
    for filename, mask in masks.items():
        if np.any(mask):
            write_point_cloud(output_dir / filename, points[mask], colors=colors[mask])


def _write_skeleton_overlay(
    path: Path,
    original_points: np.ndarray,
    triangles: np.ndarray | None,
    primary_path: np.ndarray,
    lateral_paths: list[RootPath],
    normalization: Normalization,
    metadata: dict,
) -> dict:
    d_bar = float(metadata.get("d_bar_normalized", 0.002))
    skeleton_spacing = max(d_bar * 0.55, 0.00025)
    points = [original_points]
    colors = [np.tile(SEGMENT_COLORS["unassigned"], (len(original_points), 1))]
    root_ids = [np.full(len(original_points), -1, dtype=np.int32)]
    orders = [np.full(len(original_points), 255, dtype=np.uint8)]
    states = [np.zeros(len(original_points), dtype=np.uint8)]
    faces: list[np.ndarray] = []
    has_original_faces = triangles is not None and len(triangles) > 0
    if has_original_faces:
        faces.append(np.asarray(triangles, dtype=np.int32))
    source_spacing = d_bar * float(normalization.scale)
    source_span = float(np.max(np.ptp(np.asarray(original_points, dtype=float), axis=0)))
    tube_radius = max(0.85 * source_spacing, 0.0006 * source_span, 1e-7)
    skeleton_points: list[np.ndarray] = []
    skeleton_colors: list[np.ndarray] = []
    skeleton_root_ids: list[np.ndarray] = []
    skeleton_orders: list[np.ndarray] = []
    skeleton_states: list[np.ndarray] = []
    skeleton_faces: list[np.ndarray] = []

    primary_original = normalization.inverse_points(resample_polyline(primary_path, skeleton_spacing))
    primary_vertices, primary_faces = _polyline_tube_mesh(primary_original, tube_radius)
    if len(primary_vertices):
        skeleton_points.append(primary_vertices)
        skeleton_faces.append(primary_faces)
        skeleton_colors.append(
            np.tile(SEGMENT_COLORS["primary"], (len(primary_vertices), 1))
        )
        skeleton_root_ids.append(np.zeros(len(primary_vertices), dtype=np.int32))
        skeleton_orders.append(np.zeros(len(primary_vertices), dtype=np.uint8))
        skeleton_states.append(np.ones(len(primary_vertices), dtype=np.uint8))
    for label, lateral in enumerate(lateral_paths, start=1):
        lateral_original = normalization.inverse_points(resample_polyline(lateral.points, skeleton_spacing))
        lateral_vertices, lateral_faces = _polyline_tube_mesh(lateral_original, tube_radius)
        if not len(lateral_vertices):
            continue
        offset = sum(len(values) for values in skeleton_points)
        skeleton_points.append(lateral_vertices)
        skeleton_faces.append(lateral_faces + offset)
        skeleton_colors.append(
            np.tile(order_color(lateral.order), (len(lateral_vertices), 1))
        )
        skeleton_root_ids.append(
            np.full(len(lateral_vertices), label, dtype=np.int32)
        )
        skeleton_orders.append(
            np.full(len(lateral_vertices), lateral.order, dtype=np.uint8)
        )
        skeleton_states.append(np.ones(len(lateral_vertices), dtype=np.uint8))

    skeleton_shift_x = 0.0
    comparison_gap = 0.0
    skeleton_vertex_count = 0
    if skeleton_points:
        skeleton_geometry = np.vstack(skeleton_points)
        skeleton_vertex_count = len(skeleton_geometry)
        coordinate_magnitude = max(
            abs(float(np.max(original_points[:, 0]))),
            abs(float(np.min(skeleton_geometry[:, 0]))),
            source_span,
            1.0,
        )
        representable_gap = 8.0 * abs(float(np.spacing(coordinate_magnitude)))
        comparison_gap = max(
            0.10 * source_span,
            4.0 * tube_radius,
            source_spacing,
            representable_gap,
            1e-7,
        )
        original_max_x = float(np.max(original_points[:, 0]))
        skeleton_min_x = float(np.min(skeleton_geometry[:, 0]))
        skeleton_shift_x = original_max_x + comparison_gap - skeleton_min_x
        skeleton_geometry = skeleton_geometry.copy()
        skeleton_geometry[:, 0] += skeleton_shift_x

        original_vertex_count = len(original_points)
        points.append(skeleton_geometry)
        colors.append(np.vstack(skeleton_colors))
        root_ids.append(np.concatenate(skeleton_root_ids))
        orders.append(np.concatenate(skeleton_orders))
        states.append(np.concatenate(skeleton_states))
        if has_original_faces:
            faces.append(np.vstack(skeleton_faces) + original_vertex_count)

    write_labeled_ply(
        path,
        np.vstack(points),
        triangles=np.vstack(faces) if faces else None,
        colors=np.vstack(colors),
        root_ids=np.concatenate(root_ids),
        root_orders=np.concatenate(orders),
        assignment_states=np.concatenate(states),
    )
    return {
        "arrangement": "side_by_side_x",
        "original_structure_side": "left",
        "skeleton_side": "right",
        "coordinate_unit": "mesh_unit",
        "original_translation": [0.0, 0.0, 0.0],
        "skeleton_translation": [float(skeleton_shift_x), 0.0, 0.0],
        "minimum_gap": float(comparison_gap),
        "original_vertex_count": int(len(original_points)),
        "skeleton_vertex_count": int(skeleton_vertex_count),
        "original_face_count": int(len(triangles)) if has_original_faces else 0,
        "skeleton_face_count": (
            int(sum(len(values) for values in skeleton_faces))
            if has_original_faces
            else 0
        ),
        "geometry_representation": (
            "mesh_with_skeleton_tube_faces"
            if has_original_faces
            else "point_cloud_vertices"
        ),
        "note": (
            "The original structure retains source coordinates; the skeleton "
            "translation is visualization-only."
        ),
    }


def _polyline_tube_mesh(
    polyline: np.ndarray,
    radius: float,
    *,
    sides: int = 6,
) -> tuple[np.ndarray, np.ndarray]:
    """Create capped low-poly cylinders so skeletons render in mesh viewers."""

    polyline = np.asarray(polyline, dtype=float)
    vertices: list[np.ndarray] = []
    faces: list[np.ndarray] = []
    angles = np.linspace(0.0, 2.0 * np.pi, max(3, int(sides)), endpoint=False)
    circle_cos = np.cos(angles)
    circle_sin = np.sin(angles)
    side_count = len(angles)
    for start, end in zip(polyline[:-1], polyline[1:]):
        direction = end - start
        length = float(np.linalg.norm(direction))
        if length <= 1e-12:
            continue
        direction /= length
        helper = np.array([1.0, 0.0, 0.0])
        if abs(float(np.dot(direction, helper))) > 0.85:
            helper = np.array([0.0, 1.0, 0.0])
        basis_u = np.cross(direction, helper)
        basis_u /= max(float(np.linalg.norm(basis_u)), 1e-12)
        basis_v = np.cross(direction, basis_u)
        ring_offsets = float(radius) * (
            circle_cos[:, None] * basis_u[None, :]
            + circle_sin[:, None] * basis_v[None, :]
        )
        segment_vertices = np.vstack(
            [start + ring_offsets, end + ring_offsets, start[None, :], end[None, :]]
        )
        base = len(vertices) * (2 * side_count + 2)
        segment_faces: list[list[int]] = []
        start_center = base + 2 * side_count
        end_center = start_center + 1
        for index in range(side_count):
            following = (index + 1) % side_count
            a, b = base + index, base + following
            c, d = base + side_count + index, base + side_count + following
            segment_faces.extend(
                [
                    [a, c, b],
                    [b, c, d],
                    [start_center, b, a],
                    [end_center, c, d],
                ]
            )
        vertices.append(segment_vertices)
        faces.append(np.asarray(segment_faces, dtype=np.int32))
    if not vertices:
        return np.empty((0, 3), dtype=float), np.empty((0, 3), dtype=np.int32)
    return np.vstack(vertices), np.vstack(faces)


def _json_safe(value):
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, np.ndarray):
        return _json_safe(value.tolist())
    if isinstance(value, (np.integer, np.floating)):
        return _json_safe(value.item())
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value
