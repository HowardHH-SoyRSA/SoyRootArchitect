from __future__ import annotations

import copy
import json
import math
from pathlib import Path
import xml.etree.ElementTree as ET

import numpy as np
from openpyxl import load_workbook
import pandas as pd
from PIL import Image
import pytest

from soyrootbio.export import (
    SEGMENT_COLORS,
    _polyline_tube_mesh,
    _write_skeleton_overlay,
    write_rsml,
)
from soyrootbio.geometry import resample_polyline
from soyrootbio.io import write_labeled_ply
from soyrootbio.pipeline import PipelineConfig, PipelineResult, run_pipeline
from soyrootbio.synthetic import write_synthetic_dataset
from soyrootbio.topology import apply_hierarchy_corrections
from soyrootbio.types import Normalization, RootPath


ANGLE_FIGURES = (
    "tip_gravity_front_view_600dpi.png",
    "tip_start_gravity_front_view_600dpi.png",
    "tip_primary_front_view_600dpi.png",
)

TRAIT_CSV_COLUMNS = {
    "system_summary.csv": {"trait", "value"},
    "root_lengths.csv": {"root_id", "length", "chord_length", "length_unit"},
    "lateral_counts_by_order.csv": {"root_order", "lateral_root_count"},
    "root_angles.csv": {
        "root_id",
        "tip_gravity_angle_deg",
        "tip_start_gravity_angle_deg",
        "tip_primary_angle_deg",
    },
    "root_tortuosity.csv": {"root_id", "tortuosity"},
    "root_surface_area.csv": {"root_id", "surface_area", "area_unit"},
    "root_volume.csv": {"root_id", "volume", "volume_unit"},
    "root_diameter.csv": {"root_id", "mean_diameter", "length_unit"},
    "angle_vectors.csv": {
        "root_id",
        "tip_gravity_angle_deg",
        "tip_start_gravity_angle_deg",
        "tip_primary_angle_deg",
        "base_vector_start_x",
        "base_vector_start_y",
        "base_vector_start_z",
        "base_vector_end_x",
        "base_vector_end_y",
        "base_vector_end_z",
        "base_vector_dx",
        "base_vector_dy",
        "base_vector_dz",
        "base_vector_requested_window",
        "base_vector_arc_window",
        "base_vector_window_unit",
    },
    "root_topology.csv": {"root_id", "parent_id", "root_order"},
    "root_qc.csv": {"root_id", "confidence", "qc_flags"},
}


def test_root_order_color_codes_match_export_contract() -> None:
    np.testing.assert_array_equal(
        np.round(SEGMENT_COLORS["order_1"] * 255).astype(np.uint8),
        np.array([255, 0, 255], dtype=np.uint8),
    )
    np.testing.assert_array_equal(
        np.round(SEGMENT_COLORS["order_2"] * 255).astype(np.uint8),
        np.array([0, 158, 115], dtype=np.uint8),
    )


@pytest.fixture(scope="module")
def synthetic_export(
    tmp_path_factory: pytest.TempPathFactory,
) -> tuple[Path, Path, PipelineResult]:
    work_dir = tmp_path_factory.mktemp("synthetic_export_contract")
    points_path, endpoint_path = write_synthetic_dataset(
        work_dir / "synthetic_root.csv",
        primary_points=150,
        lateral_points=70,
        lateral_count=3,
        noise=0.002,
        seed=11,
    )
    output_dir = work_dir / "outputs"
    result = run_pipeline(
        PipelineConfig(
            input_path=points_path,
            output_dir=output_dir,
            endpoint_file=endpoint_path,
            lateral_max_paths=6,
            max_root_order=2,
            random_seed=11,
        )
    )
    assert result.lateral_paths, "Synthetic fixture must exercise lateral-root exports."
    return points_path, output_dir, result


def _root_order(element: ET.Element) -> int:
    property_element = element.find("./properties/root_order")
    assert property_element is not None
    return int(property_element.attrib["value"])


def _ply_header(path: Path) -> str:
    with path.open("rb") as handle:
        header = bytearray()
        while b"end_header\n" not in header:
            chunk = handle.read(256)
            if not chunk:
                raise AssertionError(f"PLY header terminator is missing from {path}")
            header.extend(chunk)
    return bytes(header).split(b"end_header\n", 1)[0].decode("ascii") + "end_header\n"


def _read_labeled_ply(
    path: Path,
) -> tuple[np.ndarray, np.ndarray, str]:
    payload = path.read_bytes()
    marker = b"end_header\n"
    vertex_offset = payload.index(marker) + len(marker)
    header = payload[:vertex_offset].decode("ascii")
    element_counts = {
        parts[1]: int(parts[2])
        for line in header.splitlines()
        if len(parts := line.split()) == 3 and parts[0] == "element"
    }
    vertex_dtype = np.dtype(
        [
            ("x", "<f8"),
            ("y", "<f8"),
            ("z", "<f8"),
            ("red", "u1"),
            ("green", "u1"),
            ("blue", "u1"),
            ("root_id", "<i4"),
            ("root_order", "u1"),
            ("assignment_state", "u1"),
        ]
    )
    vertices = np.frombuffer(
        payload,
        dtype=vertex_dtype,
        count=element_counts["vertex"],
        offset=vertex_offset,
    ).copy()
    face_offset = vertex_offset + vertices.nbytes
    face_dtype = np.dtype([("count", "u1"), ("vertices", "<i4", (3,))])
    face_records = np.frombuffer(
        payload,
        dtype=face_dtype,
        count=element_counts["face"],
        offset=face_offset,
    )
    assert np.all(face_records["count"] == 3)
    return vertices, face_records["vertices"].copy(), header


def test_rsml_parses_and_preserves_parent_child_order(
    synthetic_export: tuple[Path, Path, PipelineResult],
    tmp_path: Path,
) -> None:
    _, output_dir, _ = synthetic_export
    pipeline_tree = ET.parse(output_dir / "root_system.rsml")
    assert pipeline_tree.getroot().tag == "rsml"
    assert pipeline_tree.findtext("./metadata/unit") == "mesh_unit"

    plant = pipeline_tree.find("./scene/plant")
    assert plant is not None
    root_elements = {element.attrib["id"]: element for element in plant.findall(".//root")}
    hierarchy = json.loads((output_dir / "root_hierarchy.json").read_text(encoding="utf-8"))
    detected_laterals = [root for root in hierarchy["roots"] if root["root_id"] != "primary"]
    assert len(detected_laterals) == 3
    assert {root["root_order"] for root in detected_laterals} == {1}
    assert {root["parent_id"] for root in detected_laterals} == {"primary"}
    for root in hierarchy["roots"]:
        element = root_elements[root["root_id"]]
        assert _root_order(element) == root["root_order"]
        if root["parent_id"] is not None:
            parent_element = root_elements[root["parent_id"]]
            assert element in parent_element.findall("./root")

    primary = np.array([[0.0, 0.0, 1.0], [0.0, 0.0, 0.0]])
    first_order = RootPath(
        root_id="order1_001",
        parent_id="primary",
        order=1,
        confidence=0.9,
        points=np.array([[0.0, 0.0, 0.7], [0.3, 0.0, 0.5]]),
    )
    second_order = RootPath(
        root_id="order2_001",
        parent_id="order1_001",
        order=2,
        confidence=0.8,
        points=np.array([[0.15, 0.0, 0.6], [0.2, 0.2, 0.45]]),
    )
    nested_traits = pd.DataFrame(
        [
            {"root_id": "primary", "length": 1.0, "mean_diameter": 0.1, "tortuosity": 1.0},
            {"root_id": "order1_001", "length": 0.4, "mean_diameter": 0.05, "tortuosity": 1.1},
            {"root_id": "order2_001", "length": 0.25, "mean_diameter": 0.03, "tortuosity": 1.2},
        ]
    )
    nested_path = write_rsml(
        tmp_path / "nested.rsml",
        primary,
        [second_order, first_order],
        nested_traits,
        {"source": "nested-fixture.csv"},
    )
    nested_tree = ET.parse(nested_path)
    assert nested_tree.findtext("./metadata/unit") == "mesh_unit"
    nested_primary = nested_tree.find("./scene/plant/root[@id='primary']")
    assert nested_primary is not None and _root_order(nested_primary) == 0
    nested_primary_points = nested_primary.findall("./geometry/polyline/point")
    np.testing.assert_allclose(
        [[float(point.attrib[axis]) for axis in ("x", "y", "z")] for point in nested_primary_points],
        primary,
    )
    nested_first = nested_primary.find("./root[@id='order1_001']")
    assert nested_first is not None and _root_order(nested_first) == 1
    nested_second = nested_first.find("./root[@id='order2_001']")
    assert nested_second is not None and _root_order(nested_second) == 2
    assert nested_primary.find("./root[@id='order2_001']") is None


def test_trait_workbook_and_per_trait_csv_contract(
    synthetic_export: tuple[Path, Path, PipelineResult],
) -> None:
    _, output_dir, _ = synthetic_export
    workbook = output_dir / "traits.xlsx"
    expected_sheet_columns = {
        "Root traits": {"root_id", "root_order"},
        "System summary": {"trait", "value"},
        "Length": {"root_id", "length", "length_unit"},
        "Counts by order": {"root_order", "lateral_root_count"},
        "Angles": {
            "tip_gravity_angle_deg",
            "tip_start_gravity_angle_deg",
            "tip_primary_angle_deg",
        },
        "Tortuosity": {"root_id", "tortuosity"},
        "Surface area": {"root_id", "surface_area", "area_unit"},
        "Volume": {"root_id", "volume", "volume_unit"},
        "Diameter": {"root_id", "mean_diameter", "length_unit"},
        "Vectors": {
            "root_id",
            "root_start_x",
            "root_tip_x",
            "base_vector_start_x",
            "base_vector_end_x",
            "base_vector_dx",
            "base_vector_arc_window",
            "base_vector_window_unit",
        },
        "Topology": {"root_id", "parent_id", "root_order"},
        "QC": {"root_id", "confidence", "qc_flags"},
        "Label map": {"numeric_label", "root_id", "root_order"},
    }
    with pd.ExcelFile(workbook, engine="openpyxl") as excel_file:
        assert set(expected_sheet_columns).issubset(excel_file.sheet_names)
        for sheet_name, expected_columns in expected_sheet_columns.items():
            frame = pd.read_excel(excel_file, sheet_name=sheet_name)
            assert expected_columns.issubset(frame.columns), sheet_name

    styled_workbook = load_workbook(workbook, read_only=False, data_only=True)
    for worksheet in styled_workbook.worksheets:
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref == worksheet.dimensions
        assert worksheet["A1"].font.bold
        assert worksheet["A1"].fill.fgColor.rgb.endswith("1F4E78")
        assert worksheet.column_dimensions["A"].width >= 10
    root_traits_sheet = styled_workbook["Root traits"]
    length_column = next(cell.column for cell in root_traits_sheet[1] if cell.value == "length")
    assert root_traits_sheet.cell(row=2, column=length_column).number_format == "0.000000"

    assert (output_dir / "root_traits.csv").stat().st_size > 0
    for filename, expected_columns in TRAIT_CSV_COLUMNS.items():
        path = output_dir / "csv" / filename
        assert path.is_file() and path.stat().st_size > 0, filename
        frame = pd.read_csv(path)
        assert expected_columns.issubset(frame.columns), filename

    summary = pd.read_csv(output_dir / "csv" / "system_summary.csv").set_index("trait")["value"]
    assert summary["root_system_surface_area_method"] == "sum_per_root_surface_estimates"
    assert summary["length_unit"] == "mesh_unit"
    assert summary["area_unit"] == "mesh_unit^2"
    assert summary["volume_unit"] == "mesh_unit^3"
    assert float(summary["unassigned_vertex_fraction"]) >= 0.0


def test_primary_qc_provenance_and_skeleton_coordinate_contract(
    synthetic_export: tuple[Path, Path, PipelineResult],
) -> None:
    _, output_dir, result = synthetic_export
    traits = pd.read_csv(output_dir / "root_traits.csv")
    topology = pd.read_csv(output_dir / "csv" / "root_topology.csv")
    hierarchy = json.loads((output_dir / "root_hierarchy.json").read_text(encoding="utf-8"))

    trait_primary = traits.loc[traits["root_id"] == "primary"].iloc[0]
    topology_primary = topology.loc[topology["root_id"] == "primary"].iloc[0]
    hierarchy_primary = next(root for root in hierarchy["roots"] if root["root_id"] == "primary")
    expected_flags = [] if pd.isna(trait_primary["qc_flags"]) else str(trait_primary["qc_flags"]).split(";")
    topology_flags = [] if pd.isna(topology_primary["qc_flags"]) else str(topology_primary["qc_flags"]).split(";")

    assert topology_primary["confidence"] == pytest.approx(trait_primary["confidence"])
    assert hierarchy_primary["confidence"] == pytest.approx(trait_primary["confidence"])
    assert topology_flags == expected_flags
    assert hierarchy_primary["qc_flags"] == expected_flags
    assert set(traits["coordinate_unit"]) == {"mesh_unit"}
    assert set(topology["coordinate_unit"]) == {"mesh_unit"}
    assert hierarchy["coordinate_unit"] == "mesh_unit"
    assert "coordinate_to_mm_scale" not in hierarchy
    assert not any("mm" in column.lower() for column in traits.columns)
    assert not any("mm" in column.lower() for column in topology.columns)

    primary_skeleton = pd.read_csv(output_dir / "primary_skeleton.csv")
    lateral_skeletons = pd.read_csv(output_dir / "lateral_skeletons.csv")
    required = {
        "x",
        "y",
        "z",
        "coordinate_unit",
    }
    for skeleton in (primary_skeleton, lateral_skeletons):
        assert required.issubset(skeleton.columns)
        assert set(skeleton["coordinate_unit"]) == {"mesh_unit"}
        assert not any("mm" in column.lower() for column in skeleton.columns)

    np.testing.assert_allclose(
        primary_skeleton[["x", "y", "z"]].to_numpy(dtype=float),
        result.normalization.inverse_points(result.primary_path),
    )


def test_labeled_ply_headers_include_labels_and_mesh_faces(
    synthetic_export: tuple[Path, Path, PipelineResult],
    tmp_path: Path,
) -> None:
    _, output_dir, _ = synthetic_export
    segmented_header = _ply_header(output_dir / "segmented_root_structure.ply")
    assert "property int root_id\n" in segmented_header
    assert "property uchar root_order\n" in segmented_header
    assert "element face 0\n" in segmented_header

    mesh_path = tmp_path / "labeled_mesh.ply"
    write_labeled_ply(
        mesh_path,
        np.array(
            [
                [0.0, 0.0, 0.0],
                [1.0, 0.0, 0.0],
                [1.0, 1.0, 0.0],
                [0.0, 1.0, 0.0],
            ]
        ),
        triangles=np.array([[0, 1, 2], [0, 2, 3]]),
        root_ids=np.array([0, 0, 1, 1]),
        root_orders=np.array([0, 0, 1, 1]),
        assignment_states=np.ones(4, dtype=np.uint8),
    )
    mesh_header = _ply_header(mesh_path)
    assert "element vertex 4\n" in mesh_header
    assert "property int root_id\n" in mesh_header
    assert "property uchar root_order\n" in mesh_header
    assert "element face 2\n" in mesh_header
    assert "property list uchar int vertex_indices\n" in mesh_header


def test_skeleton_original_overlay_is_a_side_by_side_comparison(
    tmp_path: Path,
) -> None:
    original = np.array(
        [
            [0.0, 0.0, 0.0],
            [1.0, 0.0, 0.0],
            [1.0, 0.0, 2.0],
            [0.0, 0.0, 2.0],
        ]
    )
    triangles = np.array([[0, 1, 2], [0, 2, 3]], dtype=np.int32)
    primary = np.array([[0.5, 0.0, 2.0], [0.5, 0.0, 0.0]])
    lateral = RootPath(
        root_id="root-o1-001",
        parent_id="primary",
        order=1,
        points=np.array([[0.5, 0.0, 1.2], [0.9, 0.0, 0.6]]),
    )
    second_order = RootPath(
        root_id="root-o2-001",
        parent_id="root-o1-001",
        order=2,
        points=np.array([[0.9, 0.0, 0.6], [0.8, 0.2, 0.2]]),
    )
    output = tmp_path / "skeleton_original_overlay.ply"

    layout = _write_skeleton_overlay(
        output,
        original,
        triangles,
        primary,
        [lateral, second_order],
        Normalization(minimum=np.zeros(3), scale=1.0),
        {"d_bar_normalized": 0.1},
    )

    records, faces, header = _read_labeled_ply(output)
    xyz = np.column_stack([records["x"], records["y"], records["z"]])
    original_mask = records["root_id"] == -1
    skeleton_mask = records["root_id"] >= 0
    assert np.count_nonzero(original_mask) == len(original)
    np.testing.assert_array_equal(xyz[original_mask], original)
    assert np.any(skeleton_mask)
    actual_gap = float(
        np.min(xyz[skeleton_mask, 0]) - np.max(xyz[original_mask, 0])
    )
    assert actual_gap == pytest.approx(layout["minimum_gap"])
    assert actual_gap > 0.0
    assert layout["arrangement"] == "side_by_side_x"
    assert layout["original_structure_side"] == "left"
    assert layout["skeleton_side"] == "right"
    assert layout["original_translation"] == [0.0, 0.0, 0.0]
    assert layout["skeleton_translation"][0] > 0.0
    assert layout["original_vertex_count"] == len(original)
    assert layout["skeleton_vertex_count"] == np.count_nonzero(skeleton_mask)
    assert layout["original_face_count"] == len(triangles)
    assert layout["geometry_representation"] == "mesh_with_skeleton_tube_faces"

    expected_gray = np.round(SEGMENT_COLORS["unassigned"] * 255).astype(np.uint8)
    stored_colors = np.column_stack(
        [records["red"], records["green"], records["blue"]]
    )
    np.testing.assert_array_equal(
        stored_colors[original_mask],
        np.tile(expected_gray, (len(original), 1)),
    )
    assert set(records["root_id"][skeleton_mask]) == {0, 1, 2}
    assert set(records["root_order"][skeleton_mask]) == {0, 1, 2}
    assert set(records["assignment_state"][skeleton_mask]) == {1}
    for root_id, expected_color in (
        (0, SEGMENT_COLORS["primary"]),
        (1, SEGMENT_COLORS["order_1"]),
        (2, SEGMENT_COLORS["order_2"]),
    ):
        mask = records["root_id"] == root_id
        expected_bytes = np.round(expected_color * 255).astype(np.uint8)
        np.testing.assert_array_equal(
            stored_colors[mask],
            np.tile(expected_bytes, (np.count_nonzero(mask), 1)),
        )

    skeleton_spacing = 0.055
    tube_radius = 0.085
    expected_tubes = {
        0: _polyline_tube_mesh(
            resample_polyline(primary, skeleton_spacing),
            tube_radius,
        ),
        1: _polyline_tube_mesh(
            resample_polyline(lateral.points, skeleton_spacing),
            tube_radius,
        ),
        2: _polyline_tube_mesh(
            resample_polyline(second_order.points, skeleton_spacing),
            tube_radius,
        ),
    }
    for root_id, (expected_vertices, _) in expected_tubes.items():
        actual_vertices = xyz[records["root_id"] == root_id]
        np.testing.assert_allclose(actual_vertices[:, 1:], expected_vertices[:, 1:])
        np.testing.assert_allclose(
            actual_vertices[:, 0] - expected_vertices[:, 0],
            layout["skeleton_translation"][0],
        )

    np.testing.assert_array_equal(faces[: len(triangles)], triangles)
    assert np.all(faces[len(triangles) :] >= len(original))
    assert len(faces) == len(triangles) + sum(
        len(expected_faces) for _, expected_faces in expected_tubes.values()
    )
    assert layout["skeleton_face_count"] == len(faces) - len(triangles)
    assert f"element vertex {len(records)}\n" in header
    assert f"element face {len(faces)}\n" in header


def test_pipeline_metadata_records_side_by_side_skeleton_layout(
    synthetic_export: tuple[Path, Path, PipelineResult],
) -> None:
    _, output_dir, _ = synthetic_export
    metadata = json.loads((output_dir / "metadata.json").read_text(encoding="utf-8"))
    layout = metadata["skeleton_original_overlay_layout"]
    records, faces, header = _read_labeled_ply(
        output_dir / "skeleton_original_overlay.ply"
    )
    xyz = np.column_stack([records["x"], records["y"], records["z"]])
    original_count = int(layout["original_vertex_count"])

    assert layout["arrangement"] == "side_by_side_x"
    assert original_count == metadata["source_geometry"]["full_point_count"]
    assert layout["skeleton_vertex_count"] == len(records) - original_count
    assert np.max(xyz[:original_count, 0]) < np.min(xyz[original_count:, 0])
    assert layout["geometry_representation"] == "point_cloud_vertices"
    assert layout["original_face_count"] == 0
    assert layout["skeleton_face_count"] == 0
    assert len(faces) == 0
    assert "element face 0\n" in header


def test_angle_figures_are_pngs_at_approximately_600_dpi(
    synthetic_export: tuple[Path, Path, PipelineResult],
) -> None:
    _, output_dir, _ = synthetic_export
    assert not (output_dir / "tip_angles_front_view_600dpi.png").exists()
    for filename in ANGLE_FIGURES:
        path = output_dir / filename
        assert path.is_file() and path.stat().st_size > 0
        with Image.open(path) as image:
            assert image.format == "PNG"
            dpi = image.info.get("dpi")
            assert dpi is not None and len(dpi) == 2
            assert dpi[0] == pytest.approx(600.0, abs=2.0)
            assert dpi[1] == pytest.approx(600.0, abs=2.0)


def test_hierarchy_is_editable_and_metadata_records_provenance_and_timings(
    synthetic_export: tuple[Path, Path, PipelineResult],
    tmp_path: Path,
) -> None:
    points_path, output_dir, result = synthetic_export
    hierarchy_path = output_dir / "root_hierarchy.json"
    hierarchy = json.loads(hierarchy_path.read_text(encoding="utf-8"))
    assert hierarchy["schema"] == "soyrootbio.root-hierarchy/v1"
    assert all(term in hierarchy["instructions"] for term in ("parent_id", "valid", "polyline"))

    roots = hierarchy["roots"]
    by_id = {root["root_id"]: root for root in roots}
    assert len(by_id) == len(roots)
    assert by_id["primary"]["parent_id"] is None
    assert by_id["primary"]["root_order"] == 0
    for root in roots:
        polyline = np.asarray(root["polyline"], dtype=float)
        assert polyline.ndim == 2 and polyline.shape[1] == 3 and len(polyline) >= 2
        assert np.all(np.isfinite(polyline))
        if root["parent_id"] is not None:
            parent = by_id[root["parent_id"]]
            assert root["root_order"] == parent["root_order"] + 1

    parent_ids = {root["parent_id"] for root in roots if root["parent_id"] is not None}
    lateral_to_remove = next(
        root
        for root in roots
        if root["root_id"] != "primary" and root["root_id"] not in parent_ids
    )
    correction_path = tmp_path / "hierarchy_edit.json"
    correction_path.write_text(
        json.dumps(
            {
                "schema": hierarchy["schema"],
                # Old correction files may retain this now-ignored field.
                "coordinate_to_mm_scale": 0.125,
                "roots": [{"root_id": lateral_to_remove["root_id"], "valid": False}],
            }
        ),
        encoding="utf-8",
    )
    corrected = apply_hierarchy_corrections(
        result.primary_path.copy(),
        copy.deepcopy(result.lateral_paths),
        correction_path,
    )
    assert len(corrected) == len(result.lateral_paths) - 1
    assert lateral_to_remove["root_id"] not in {root.root_id for root in corrected}

    metadata_text = (output_dir / "metadata.json").read_text(encoding="utf-8")
    metadata = json.loads(
        metadata_text,
        parse_constant=lambda token: pytest.fail(f"Non-finite JSON constant in metadata: {token}"),
    )
    assert metadata["source"] == str(points_path)
    assert metadata["algorithm_reference"].startswith("Zhou et al. 2025")
    assert metadata["config"]["input_path"] == str(points_path)
    assert metadata["config"]["endpoint_file"].endswith("synthetic_root_endpoints.csv")
    assert metadata["primary_detection_method"] == "manual_endpoints_with_optional_sections"
    assert metadata["gravity_vector"] == [0.0, 0.0, -1.0]
    assert metadata["output_length_unit"] == "mesh_unit"
    assert metadata["source_geometry"]["geometry_kind"] == "point_cloud"
    assert metadata["point_count"] == result.point_count
    assert metadata["selected_lateral_count"] == len(result.lateral_paths)
    assignment = metadata["point_assignment"]
    assert assignment["total_vertex_count"] == len(result.full_root_labels)
    assert assignment["unassigned_reason_counts"]["above_selected_base"] == int(
        np.count_nonzero(result.full_above_base_mask)
    )
    assert sum(assignment["unassigned_reason_counts"].values()) == assignment[
        "unassigned_vertex_count"
    ]

    expected_stages = {
        "load_geometry",
        "normalization",
        "primary_detection",
        "primary_segmentation",
        "lateral_tracing",
        "topology_repair",
        "point_assignment",
        "trait_measurement",
        "validation_figures",
        "export",
        "total",
    }
    timings = metadata["stage_timings_seconds"]
    assert expected_stages.issubset(timings)
    assert all(
        isinstance(timings[stage], (int, float))
        and math.isfinite(timings[stage])
        and timings[stage] >= 0.0
        for stage in expected_stages
    )
    assert timings["total"] > 0.0
    assert timings["total"] >= max(timings[stage] for stage in expected_stages - {"total"})

    recorded_outputs = set(metadata["outputs"])
    assert {
        "root_system.rsml",
        "root_hierarchy.json",
        "traits.xlsx",
        "segmented_root_structure.ply",
        *(ANGLE_FIGURES),
    }.issubset(recorded_outputs)
    actual_outputs = {
        str(path.relative_to(output_dir)).replace("\\", "/")
        for path in output_dir.rglob("*")
        if path.is_file() and path.name != "metadata.json"
    }
    assert recorded_outputs == actual_outputs


def test_unchanged_exported_hierarchy_round_trip_is_not_reported_as_manual_edit(
    synthetic_export: tuple[Path, Path, PipelineResult],
    tmp_path: Path,
) -> None:
    points_path, output_dir, _ = synthetic_export
    original_metadata = json.loads((output_dir / "metadata.json").read_text(encoding="utf-8"))
    correction_file = output_dir / "root_hierarchy.json"
    round_trip_output = tmp_path / "round-trip"

    rerun = run_pipeline(
        PipelineConfig(
            input_path=points_path,
            output_dir=round_trip_output,
            endpoint_file=Path(original_metadata["config"]["endpoint_file"]),
            correction_file=correction_file,
            lateral_max_paths=6,
            max_root_order=2,
            random_seed=11,
        )
    )

    correction_metadata = json.loads(
        (round_trip_output / "metadata.json").read_text(encoding="utf-8")
    )["hierarchy_correction"]
    exported_ids = {
        root["root_id"]
        for root in json.loads(correction_file.read_text(encoding="utf-8"))["roots"]
        if root["root_id"] != "primary"
    }
    assert {root.root_id for root in rerun.lateral_paths} == exported_ids
    assert correction_metadata["applied"] is True
    assert correction_metadata["removed_root_ids"] == []
    assert correction_metadata["manually_changed_root_ids"] == []
